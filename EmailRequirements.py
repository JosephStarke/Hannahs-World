# -*- coding: utf-8 -*-
"""
Created on Mon Jun 13 13:42:34 2022

@author: Joseph Starke, jwstarke00@gmail.com

Creation Steps
0. Recreate .XLSX file in a good format
1. Open .XLSX file
2. Dump file
3. Make people objects and fill with the information
3.5. Map info from data index numbers to the type of data
4. Itterate through each person and see if they are missing objects or not, if they are print the correct message
5. Set up sending emails
6. Make a test file with test email info (one with no people, 1 person, multiple people, all need to send, none need to send, some need to send)
7. Run program with the test files
8. Get correct Email from Hannah, set up a dummy email to see if stuff sends correctly every email.
9. Test with real info
10. Turn it into an executable and send it to Hannah

Executable Creation Command
- auto-py-to-exe
- Run it as command line

STATUS

STATUS: Working as desried, need email column.
BUGS: None known.

Installs
"""
# Modules for text change
import datetime
# Modules for Exel
import openpyxl
import os
import sys
# Modules for emails
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib


###VARIABLES
#Time
time = datetime.datetime.now()
#Email
subject = 'Early Childhood Missing Documents'
message = '' #leave blank, remade later
sender = 'hwcacademy@gmail.com'
appPassword = 'kzhhzcbmrdagngvs'

#Read File
people = []
###CODE

#Connection to email server
print("Openning email connection...")
server = smtplib.SMTP_SSL('smtp.gmail.com', 465) #Number comes from: Outlook desktop app, File -> Account Settings -> Server Settings -> Outgoing Mail
server.login(sender, appPassword) #Email address and a read into a file containing the email's password (more secure), it should be the app password


#message generation
def generate_email(lastName, firstName, missingInfo, reciever):
    message = ('Good afternoon ' + firstName + ' ' + lastName + ','
               '\n'
               '\n' 
               'Our records indicate you are missing: '
               '\n'
               '\n' 
               + missingInfo +
               '\n'
               '\n' 
               'Please submit it to your director within the next three days of operation. Your cooperation is greatly appreciated!'
               '\n'
               '\n' 
               'Best regards, '
               '\n'
               'Your Holy Word Christian Academy Administration'
               )
    
    print (message)
    
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg.attach(MIMEText(message))
    
    #send email
    server.sendmail(sender,
                  reciever, 
                  msg.as_string())
    
#person class to store data per person
class Person:
    def __init__(self, data):
        #CONSTRUCTOR INFO
        #Needed
        self.data = data
        
#dictionaries
data_map = {
    #Required
    0:'Last Name', 
    1:'First Name',
    2:'Email',
    3:'Child Care Center Personnel Information Record', 
    4:'Personal History Statement', 
    5:'Copy of Photo Identification',
    6:'Background Check Form', 
    7:'Background Check / Fingerprinting Clearance', 
    8:'Affidavit for Applicants for Employment', 
    9:'Educational Documentation', 
    10:'CPR/First Aid Certificates',
    11:'Staff Training Record', 
    12:'Pre-Training Certificate', 
    13:'Certificates Verifying Clock Hours', 
    14:'W4 Form', 
    15:'Job Description Acknowledgement',
    16:'Employee Handbook Agreement', 
    17:'1-9 Form', 
    18:'1-2 Forms of ID Copies',
    #Not Required
    19:'TB Test', 
    20:'Direct Deposit Info',
    21:'Employee COVID Agreement', 
    22:'Doctor\'s Notes ', 
    23:'Unemployment Info', }

print("Reading file...")
# Define variable to load the wookbook
if (not os.path.isfile('Certification Checklists.xlsx')):
    print('\n[ERROR] There is no file named: Certification Checklists.xlsx')
    print('')
    print('There may be numbers automatically placed behind the name from consecutive downloads, make sure there are no numbers! OR this program may have added a timeframe to the front of the file!')
    print('\nIf this is the case, you can right click the file and select rename to name the file the the approriate name.')
    server.quit()
    
    print("")
    print("Email connection closed.")
    
    print("")
    print("Terminating program.")    
    sys.exit()
    
workbook = openpyxl.load_workbook("Certification Checklists.xlsx")

# Define variable to read the active sheet:
worksheet = workbook.active

# Iterate the loop to read the cell values
for i in range(0, worksheet.max_row):
    
    curPersonInfo = []
    
    for col in worksheet.iter_cols(1, worksheet.max_column):
        
        #Remove rows 1 and 3 cause they are uneeded
        if i == 0 or i == 2:
            break
        #Remove any Null values
        if col[i].value == None:
            break
        
        #Add the infomation to the list of information for the specific person
        curPersonInfo.append(str(col[i].value))
    
    #Make a new person object
    person = Person(curPersonInfo)
    
    #Add the person to the list of people
    people.append(person)
print("Finished reading file.")
print("")
print("Sending emails...")
  
    
  
#Check for incomplete items for each person
flaggedInfo = []

for i in range(0, len(people)):
    print('Checking person #' + str(i+1) + '...\n\n')
    for j in range(0, len(people[i].data)):
        if people[i].data[j] == 'N':
            flaggedInfo.append(j)
    
    #EMAIL PERSON HERE
    missingInfo = ''
    
    for s in range(0, len(flaggedInfo)):
        if len(flaggedInfo) > 1:
            #if not final item
            if s < len(flaggedInfo) - 1:
                missingInfo += '• ' + (data_map[flaggedInfo[s]] + ', \n')
            #if it is the final item (GENERATE EMAIL HERE)
            else:
                missingInfo += ('• ' + data_map[flaggedInfo[s]] + '.')
                #build message content for email
                generate_email(people[i].data[0], people[i].data[1], missingInfo, people[i].data[2]) #Makes this the email column
        else:
            missingInfo += (data_map[flaggedInfo[s]] + '.')
            generate_email(people[i].data[0], people[i].data[1], missingInfo, people[i].data[2])  #Makes this the email column
        
    flaggedInfo.clear()
    
print("Sending emails complete.")
print("")

worksheet = workbook.close()
server.quit()

print("Email connection closed.")
print("")
print("Renaming file...")
if (os.path.isfile('Certification Checklists.xlsx')):
    print('\n')
    print("[WARNING]: The file cannot be renamed because another file with the date in name already exists.")
else:    
    os.rename('Certification Checklists.xlsx', time.strftime("%m-%d-%Y") + ' Certification Checklists.xlsx' )
print("")
print("Terminating program.")

